"""Bulk SKU import (§98, PHASE 21).

§98's acceptance is "import dozens of SKUs and queue stably", and the two words
that shape this module are *dozens* and *stably*.

**Dozens** means the failure mode is not a crash, it is a spreadsheet where
row 34 has a blank name and row 51 has a duration of "five". So parsing and
validation are separate from execution, and the result of validation is a
*preview* (P21-T03): every row's verdict, shown before anybody commits to
generating fifty videos. A batch that validated as it ran would spend money on
thirty-three rows before reaching the bad one.

**Stably** means one tenant's import of two hundred SKUs must not fill every
queue and starve every other workspace's single video. `max_concurrency` per
batch is that bound (P21-T08), enforced by counting the batch's own running
items before dispatching more — not by a semaphore, which would not survive a
worker restart.

**Why rows become items rather than one job.** Fifty independent items can each
fail, be inspected and be retried on their own. A batch modelled as a single
job makes "row 34 failed" unanswerable and "retry row 34" impossible.
"""

from __future__ import annotations

import csv
import io
import uuid
from dataclasses import dataclass, field
from datetime import UTC, datetime
from typing import Any, Final

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from backend_core.config import Settings, get_settings
from backend_core.domain.enums import (
    TERMINAL_BATCH_ITEM,
    BatchItemStatus,
    BatchStatus,
)
from backend_core.domain.models import Batch, BatchItem, Product
from backend_core.errors import NotFoundError, ValidationError
from backend_core.observability import get_logger

logger = get_logger(__name__)

#: Rows one import may carry. §98 says "dozens"; this allows an order of
#: magnitude more and refuses the accidental 100k-row export, which would take
#: an hour to validate and produce a queue nobody wanted.
MAX_ROWS: Final[int] = 500

#: Upload ceiling for the source file. A 5 MB CSV is roughly 50k rows — well
#: past `MAX_ROWS`, so this bound only ever catches something pathological.
MAX_SOURCE_BYTES: Final[int] = 5 * 1024 * 1024

#: Columns a row must carry. Everything else is optional, because a bulk import
#: whose required set is large is one nobody can produce from their own data.
REQUIRED_COLUMNS: Final[tuple[str, ...]] = ("name", "category")

#: Every column the importer understands. An unknown column is reported rather
#: than ignored: a misspelled `catagory` that is silently dropped produces
#: fifty products with no category and no explanation.
KNOWN_COLUMNS: Final[frozenset[str]] = frozenset(
    {
        "name",
        "category",
        "brand_name",
        "sku",
        "description",
        "image_url",
        "duration_seconds",
        "aspect_ratio",
        "language",
        "target_audience",
    }
)


@dataclass(frozen=True, slots=True)
class ParsedRow:
    """One source row, with its original line number (§98, P21-T06)."""

    row_number: int
    values: dict[str, str]
    errors: tuple[str, ...] = ()

    @property
    def valid(self) -> bool:
        return not self.errors


@dataclass(frozen=True, slots=True)
class ImportPreview:
    """What an import *would* do (P21-T03)."""

    rows: tuple[ParsedRow, ...]
    unknown_columns: tuple[str, ...] = ()
    missing_columns: tuple[str, ...] = ()

    @property
    def valid_rows(self) -> tuple[ParsedRow, ...]:
        return tuple(row for row in self.rows if row.valid)

    @property
    def invalid_rows(self) -> tuple[ParsedRow, ...]:
        return tuple(row for row in self.rows if not row.valid)

    @property
    def importable(self) -> bool:
        """Whether anything can be imported at all.

        Missing required *columns* is fatal for the file; individual bad rows
        are not. A file with no `name` column cannot produce a single product,
        while a file with three bad rows out of fifty can produce forty-seven.
        """
        return not self.missing_columns and bool(self.valid_rows)


# --- parsing (P21-T01, P21-T02) --------------------------------------------


def parse_csv(data: bytes) -> ImportPreview:
    """Parse a CSV export (P21-T01).

    Decoded as UTF-8 with a BOM tolerance, because that is what Excel writes
    when a Chinese-locale user saves as CSV — and a first column named
    `\\ufeffname` is the single most common reason a real import fails.
    """
    text = _decode(data)
    reader = csv.DictReader(io.StringIO(text))
    headers = [(name or "").strip().lower() for name in (reader.fieldnames or [])]

    rows: list[ParsedRow] = []
    for offset, raw in enumerate(reader):
        # +2: the header is line 1 and spreadsheets are 1-indexed, so this is
        # the line number the user sees.
        rows.append(_validate_row(offset + 2, _normalise(raw)))
        if len(rows) >= MAX_ROWS:
            break

    return _preview(rows, headers)


def parse_xlsx(data: bytes) -> ImportPreview:
    """Parse an .xlsx workbook's first sheet (P21-T02).

    `read_only` and `data_only`: the first keeps a large sheet off the heap,
    the second returns a formula's cached *value* rather than its text — a
    column of `=CONCAT(...)` product names is otherwise imported as the
    formulas themselves.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        iterator = sheet.iter_rows(values_only=True)

        try:
            header_row = next(iterator)
        except StopIteration:
            return ImportPreview(rows=(), missing_columns=REQUIRED_COLUMNS)

        headers = [str(cell).strip().lower() if cell is not None else "" for cell in header_row]

        rows: list[ParsedRow] = []
        for offset, raw in enumerate(iterator):
            values = {
                headers[index]: _cell(value)
                for index, value in enumerate(raw)
                if index < len(headers) and headers[index]
            }
            if not any(values.values()):
                # A trailing blank row. Excel produces these constantly, and
                # reporting each as invalid would bury the real errors.
                continue
            rows.append(_validate_row(offset + 2, values))
            if len(rows) >= MAX_ROWS:
                break

        return _preview(rows, headers)
    finally:
        workbook.close()


def parse_upload(data: bytes, filename: str) -> ImportPreview:
    """Parse by extension, refusing anything else (P21-T01, P21-T02)."""
    if len(data) > MAX_SOURCE_BYTES:
        raise ValidationError(
            "That file is too large to import.",
            details={"bytes": len(data), "limit": MAX_SOURCE_BYTES},
        )

    lowered = filename.lower()
    if lowered.endswith(".csv"):
        return parse_csv(data)
    if lowered.endswith((".xlsx", ".xlsm")):
        return parse_xlsx(data)
    raise ValidationError("Import a .csv or .xlsx file.", details={"filename": filename[:120]})


def _decode(data: bytes) -> str:
    """UTF-8 with a BOM, then UTF-8, then GBK.

    GBK last because a Chinese-locale Excel still writes it by default, and a
    file that decodes as mojibake produces fifty products with unreadable
    names rather than an error anyone can act on.
    """
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        decoded = _try_decode(data, encoding)
        if decoded is not None:
            return decoded
    raise ValidationError("That file's text encoding could not be read.")


def _try_decode(data: bytes, encoding: str) -> str | None:
    try:
        return data.decode(encoding)
    except UnicodeDecodeError:
        # Expected while probing encodings — the caller tries the next one.
        return None


def _cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        # openpyxl returns every number as a float, so a duration of 15 arrives
        # as 15.0 and would be imported as the string "15.0".
        return str(int(value))
    return str(value).strip()


def _normalise(raw: dict[str | Any, str | Any]) -> dict[str, str]:
    return {
        str(key).strip().lower(): (str(value).strip() if value is not None else "")
        for key, value in raw.items()
        if key is not None
    }


def _preview(rows: list[ParsedRow], headers: list[str]) -> ImportPreview:
    present = {header for header in headers if header}
    return ImportPreview(
        rows=tuple(rows),
        unknown_columns=tuple(sorted(present - KNOWN_COLUMNS)),
        missing_columns=tuple(column for column in REQUIRED_COLUMNS if column not in present),
    )


def _validate_row(row_number: int, values: dict[str, str]) -> ParsedRow:
    """Check one row (P21-T03).

    Every problem is collected rather than raising on the first. A user fixing
    a spreadsheet wants the whole list; returning one error per round trip
    turns a five-minute fix into an afternoon.
    """
    errors: list[str] = []

    for column in REQUIRED_COLUMNS:
        if not values.get(column, "").strip():
            errors.append(f"{column} is required")

    if len(values.get("name", "")) > 200:
        errors.append("name is longer than 200 characters")
    if len(values.get("sku", "")) > 120:
        errors.append("sku is longer than 120 characters")

    duration = values.get("duration_seconds", "").strip()
    if duration:
        try:
            seconds = int(float(duration))
        except ValueError:
            errors.append(f"duration_seconds is not a number: {duration!r}")
        else:
            if not 5 <= seconds <= 180:
                errors.append("duration_seconds must be between 5 and 180")

    ratio = values.get("aspect_ratio", "").strip()
    if ratio and ratio not in {"9:16", "16:9", "1:1", "4:5"}:
        errors.append(f"aspect_ratio is not supported: {ratio!r}")

    url = values.get("image_url", "").strip()
    if url and not url.startswith(("http://", "https://")):
        # Fetched later through §61's safe downloader; a non-HTTP scheme is
        # rejected here so the reason reaches the person editing the file.
        errors.append("image_url must be an http or https URL")

    return ParsedRow(row_number=row_number, values=values, errors=tuple(errors))


# --- the batch itself (P21-T04 through T08) ---------------------------------


@dataclass(slots=True)
class BatchProgress:
    """A batch's counts, for the per-item status view (P21-T06)."""

    total: int = 0
    pending: int = 0
    running: int = 0
    completed: int = 0
    failed: int = 0
    invalid: int = 0
    by_status: dict[str, int] = field(default_factory=dict)

    @property
    def finished(self) -> bool:
        return self.pending == 0 and self.running == 0


class BatchService:
    """Import, validate, queue and retry (§98)."""

    def __init__(self, session: AsyncSession, *, settings: Settings | None = None) -> None:
        self._session = session
        self._settings = settings or get_settings()

    # -- import (P21-T01 through T04) ---------------------------------------

    async def create_from_upload(
        self,
        *,
        workspace_id: uuid.UUID,
        name: str,
        data: bytes,
        filename: str,
        template_id: uuid.UUID | None = None,
        brand_kit_id: uuid.UUID | None = None,
        created_by: uuid.UUID | None = None,
        max_concurrency: int = 3,
    ) -> tuple[Batch, ImportPreview]:
        """Parse a file into a batch of items (P21-T04).

        Invalid rows are stored as `INVALID` items rather than dropped. §98
        wants a validation preview, and a preview that omitted the bad rows
        would tell a user their file was fine while quietly importing less than
        they uploaded.
        """
        preview = parse_upload(data, filename)

        if preview.missing_columns:
            raise ValidationError(
                "That file is missing required columns.",
                details={
                    "missing": list(preview.missing_columns),
                    "required": list(REQUIRED_COLUMNS),
                },
            )
        if not preview.rows:
            raise ValidationError("That file has no rows to import.")

        batch = Batch(
            workspace_id=workspace_id,
            name=name,
            status=BatchStatus.READY if preview.valid_rows else BatchStatus.FAILED,
            template_id=template_id,
            brand_kit_id=brand_kit_id,
            source_filename=filename[:255],
            source_format="xlsx" if filename.lower().endswith((".xlsx", ".xlsm")) else "csv",
            total_items=len(preview.rows),
            created_by=created_by,
            max_concurrency=max(1, min(max_concurrency, 20)),
        )
        self._session.add(batch)
        await self._session.flush()

        for row in preview.rows:
            self._session.add(
                BatchItem(
                    workspace_id=workspace_id,
                    batch_id=batch.id,
                    row_number=row.row_number,
                    status=(BatchItemStatus.PENDING if row.valid else BatchItemStatus.INVALID),
                    source_row=dict(row.values),
                    validation_errors=list(row.errors),
                )
            )
        await self._session.flush()

        logger.info(
            "batch_imported",
            extra={
                "batch_id": str(batch.id),
                "rows": len(preview.rows),
                "valid": len(preview.valid_rows),
                "invalid": len(preview.invalid_rows),
                "unknown_columns": list(preview.unknown_columns),
            },
        )
        return batch, preview

    # -- reads (P21-T06) ----------------------------------------------------

    async def get(self, *, workspace_id: uuid.UUID, batch_id: uuid.UUID) -> Batch:
        result = await self._session.execute(
            select(Batch).where(Batch.id == batch_id, Batch.workspace_id == workspace_id)
        )
        batch = result.scalar_one_or_none()
        if batch is None:
            raise NotFoundError("Batch not found.", details={"batch_id": str(batch_id)})
        return batch

    async def list_batches(self, *, workspace_id: uuid.UUID, limit: int = 50) -> list[Batch]:
        result = await self._session.execute(
            select(Batch)
            .where(Batch.workspace_id == workspace_id)
            .order_by(Batch.created_at.desc())
            .limit(limit)
        )
        return list(result.scalars().all())

    async def items(
        self,
        *,
        workspace_id: uuid.UUID,
        batch_id: uuid.UUID,
        status: BatchItemStatus | None = None,
    ) -> list[BatchItem]:
        query = select(BatchItem).where(
            BatchItem.workspace_id == workspace_id, BatchItem.batch_id == batch_id
        )
        if status is not None:
            query = query.where(BatchItem.status == status)
        result = await self._session.execute(query.order_by(BatchItem.row_number))
        return list(result.scalars().all())

    async def progress(self, *, workspace_id: uuid.UUID, batch_id: uuid.UUID) -> BatchProgress:
        """Counts by status (P21-T06).

        Aggregated in the database rather than by loading every item: a batch
        of five hundred is exactly the case this view exists for.
        """
        result = await self._session.execute(
            select(BatchItem.status, func.count(BatchItem.id))
            .where(BatchItem.workspace_id == workspace_id, BatchItem.batch_id == batch_id)
            .group_by(BatchItem.status)
        )
        counts = {status: int(count) for status, count in result.all()}

        return BatchProgress(
            total=sum(counts.values()),
            pending=counts.get(BatchItemStatus.PENDING, 0) + counts.get(BatchItemStatus.QUEUED, 0),
            running=counts.get(BatchItemStatus.RUNNING, 0),
            completed=counts.get(BatchItemStatus.COMPLETED, 0),
            failed=counts.get(BatchItemStatus.FAILED, 0),
            invalid=counts.get(BatchItemStatus.INVALID, 0),
            by_status={status.value: count for status, count in counts.items()},
        )

    # -- queueing (P21-T05, P21-T08) ----------------------------------------

    async def claim_next(self, *, workspace_id: uuid.UUID, batch_id: uuid.UUID) -> list[BatchItem]:
        """Move as many pending items to `QUEUED` as concurrency allows.

        The concurrency bound is enforced by *counting the batch's own running
        items* rather than by a semaphore. A semaphore lives in one process and
        does not survive a restart; a count derived from the rows is correct
        after any failure, which is the property §98's "stably" needs.

        `FOR UPDATE SKIP LOCKED` on the pending rows lets two dispatchers run
        without handing the same row to both — and without either waiting on
        the other, which a plain `FOR UPDATE` would.
        """
        batch = await self.get(workspace_id=workspace_id, batch_id=batch_id)
        if batch.status not in (BatchStatus.READY, BatchStatus.RUNNING):
            return []

        in_flight = await self._session.execute(
            select(func.count(BatchItem.id)).where(
                BatchItem.batch_id == batch_id,
                BatchItem.status.in_([BatchItemStatus.QUEUED, BatchItemStatus.RUNNING]),
            )
        )
        capacity = batch.max_concurrency - int(in_flight.scalar() or 0)
        if capacity <= 0:
            return []

        result = await self._session.execute(
            select(BatchItem)
            .where(
                BatchItem.batch_id == batch_id,
                BatchItem.status == BatchItemStatus.PENDING,
            )
            .order_by(BatchItem.row_number)
            .limit(capacity)
            .with_for_update(skip_locked=True)
        )
        claimed = list(result.scalars().all())

        for item in claimed:
            item.status = BatchItemStatus.QUEUED
            item.attempts += 1

        if claimed and batch.status is BatchStatus.READY:
            batch.status = BatchStatus.RUNNING
        await self._session.flush()

        logger.info(
            "batch_items_claimed",
            extra={"batch_id": str(batch_id), "claimed": len(claimed), "capacity": capacity},
        )
        return claimed

    async def mark_running(self, item: BatchItem) -> None:
        item.status = BatchItemStatus.RUNNING
        item.started_at = datetime.now(UTC)
        await self._session.flush()

    async def mark_completed(
        self,
        item: BatchItem,
        *,
        product_id: uuid.UUID | None = None,
        project_id: uuid.UUID | None = None,
    ) -> None:
        item.status = BatchItemStatus.COMPLETED
        item.product_id = product_id
        item.project_id = project_id
        item.finished_at = datetime.now(UTC)
        item.error_code = None
        item.error_message = None
        await self._session.flush()
        await self._refresh_counts(item.batch_id)

    async def mark_failed(
        self, item: BatchItem, *, error_code: str, error_message: str = ""
    ) -> None:
        item.status = BatchItemStatus.FAILED
        item.error_code = error_code[:64]
        item.error_message = error_message[:2000] or None
        item.finished_at = datetime.now(UTC)
        await self._session.flush()
        await self._refresh_counts(item.batch_id)

    async def _refresh_counts(self, batch_id: uuid.UUID) -> None:
        """Roll item outcomes up onto the batch.

        Recomputed from the items rather than incremented, so a retry that
        turns a failure into a success corrects the total instead of leaving
        the batch permanently one short.
        """
        batch = await self._session.get(Batch, batch_id)
        if batch is None:  # pragma: no cover
            return

        result = await self._session.execute(
            select(BatchItem.status, func.count(BatchItem.id))
            .where(BatchItem.batch_id == batch_id)
            .group_by(BatchItem.status)
        )
        counts = {status: int(count) for status, count in result.all()}

        batch.completed_items = counts.get(BatchItemStatus.COMPLETED, 0)
        batch.failed_items = counts.get(BatchItemStatus.FAILED, 0)

        outstanding = sum(
            count for status, count in counts.items() if status not in TERMINAL_BATCH_ITEM
        )
        if outstanding == 0:
            if batch.failed_items and batch.completed_items:
                batch.status = BatchStatus.PARTIAL
            elif batch.failed_items:
                batch.status = BatchStatus.FAILED
            else:
                batch.status = BatchStatus.COMPLETED
        await self._session.flush()

    # -- retry (P21-T07) ----------------------------------------------------

    async def retry_failed(
        self, *, workspace_id: uuid.UUID, batch_id: uuid.UUID
    ) -> list[BatchItem]:
        """Return failed items to `PENDING` so the dispatcher picks them up.

        `INVALID` items are deliberately *not* retried: their row is wrong, and
        running it again produces the same wrong row. Fixing those means
        re-importing the corrected file.
        """
        batch = await self.get(workspace_id=workspace_id, batch_id=batch_id)
        failed = await self.items(
            workspace_id=workspace_id, batch_id=batch_id, status=BatchItemStatus.FAILED
        )
        for item in failed:
            item.status = BatchItemStatus.PENDING
            item.error_code = None
            item.error_message = None
            item.finished_at = None

        if failed:
            batch.status = BatchStatus.RUNNING
            batch.failed_items = 0
        await self._session.flush()

        logger.info("batch_retry", extra={"batch_id": str(batch_id), "retried": len(failed)})
        return failed

    async def cancel(self, *, workspace_id: uuid.UUID, batch_id: uuid.UUID) -> Batch:
        """Stop a batch. Items already running are left to finish.

        Interrupting a generation mid-flight would waste what has been paid
        for; skipping what has not started is free.
        """
        batch = await self.get(workspace_id=workspace_id, batch_id=batch_id)
        pending = await self.items(
            workspace_id=workspace_id, batch_id=batch_id, status=BatchItemStatus.PENDING
        )
        for item in pending:
            item.status = BatchItemStatus.SKIPPED
        batch.status = BatchStatus.CANCELED
        await self._session.flush()
        return batch

    # -- turning a row into a product ---------------------------------------

    async def materialise(self, item: BatchItem) -> Product:
        """Create the product a row describes (P21-T05).

        Only the product. The project, storyboard and generation are the
        pipeline's, and duplicating them here would give batch imports a second
        implementation of the same rules that could drift from the first.
        """
        row = item.source_row
        product = Product(
            workspace_id=item.workspace_id,
            name=str(row.get("name", ""))[:200],
            category=str(row.get("category", ""))[:120],
            brand_name=(str(row.get("brand_name", "")) or None),
            sku=(str(row.get("sku", "")) or None),
            description=(str(row.get("description", "")) or None),
        )
        self._session.add(product)
        await self._session.flush()

        item.product_id = product.id
        await self._session.flush()
        return product


__all__ = [
    "KNOWN_COLUMNS",
    "MAX_ROWS",
    "MAX_SOURCE_BYTES",
    "REQUIRED_COLUMNS",
    "BatchProgress",
    "BatchService",
    "ImportPreview",
    "ParsedRow",
    "parse_csv",
    "parse_upload",
    "parse_xlsx",
]
